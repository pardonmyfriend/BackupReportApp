import re
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.formatting.rule import FormulaRule
import locale


locale.setlocale(locale.LC_TIME, 'pl_PL.UTF-8')


MONTHS_MAP = {
    'stycznia': 'styczeń',
    'lutego': 'luty',
    'marca': 'marzec',
    'kwietnia': 'kwiecień',
    'maja': 'maj',
    'czerwca': 'czerwiec',
    'lipca': 'lipiec',
    'sierpnia': 'sierpień',
    'września': 'wrzesień',
    'października': 'październik',
    'listopada': 'listopad',
    'grudnia': 'grudzień'
}


def replace_months(date_str):
    for key, value in MONTHS_MAP.items():
        date_str = date_str.replace(key, value)
    return date_str


def merge_retry_rows(df):
    backup_columns = df.columns[3:]
    rows_to_remove = []

    for i, row in df.iterrows():
        backup_job = row['Backup job']

        for col1 in backup_columns:
            if pd.notna(row[col1]):
                break
            else:
                for col2 in backup_columns:
                    if pd.notna(row[col2]):
                        for j in range(i - 1, -1, -1):
                            if df.loc[j, 'Backup job'] == backup_job and pd.isna(df.loc[j, col2]):
                                df.loc[j, col2] = row[col2]
                                df.loc[j, 'Status'] = row['Status']
                                rows_to_remove.append(i)
                                break
    df = df.drop(rows_to_remove).reset_index(drop=True)
    return df


def get_backup_execution(sheet):
    backup_jobs = []
    current_job = None
    week_num = 1
    current_entry = None
    retry_num = None

    for row in sheet.iter_rows(values_only=True):
        if row[0] and "Backup job" in row[0]:
            status = row[8]
            if "Retry" in row[0]:
                match = re.search(r'Backup job: (.*?) \(Retry (\d+)\)', row[0])
                if match:
                    current_job = match.group(1)
                    retry_num = int(match.group(2))
            else:
                current_job = row[0].split("Backup job: ")[1].strip()
            current_entry = {
                'Date': None,
                'Week number': None,
                'Day of week': None,
                'Backup job': current_job,
                'Status': status
            }
        elif row[0] and re.search(r"\d{1,2}:\d{2}:\d{2}", row[0]):
            day_of_week = row[0].split(',')[0]
            date_str = replace_months(row[0].split(',')[-1].strip())
            current_entry['Date'] = datetime.strptime(date_str, '%d %B %Y %H:%M:%S').date()
            match = re.search(r'\b(\d{1,2})\b', date_str)
            if match:
                day = int(match.group(1))
            week_num = (day - 1) // 7 + 1
            current_entry['Week number'] = week_num
            current_entry['Day of week'] = day_of_week
        elif row[3] and row[2] == "Start time" and row[3] != "End time":
            start_time = row[3]
            if retry_num is None:
                current_entry['Backup'] = start_time
            else:
                current_entry[f'Backup (Retry {retry_num})'] = start_time
                retry_num = None
            backup_jobs.append(current_entry)
            current_entry = {
                'Date': None,
                'Week number': None,
                'Day of week': None,
                'Backup job': current_job,
                'Status': None
            }
    df = pd.DataFrame(backup_jobs)

    for row in range(len(df)):
        for col in df.columns[5:]:
            if pd.notna(df.loc[row, col]):
                date = pd.to_datetime(df.at[row, 'Date'])
                time = pd.to_datetime(df.at[row, col], format='%H:%M:%S').time()
                df.at[row, 'Start datetime'] = datetime.combine(date, time)
    
    df = df.sort_values(by='Start datetime').reset_index(drop=True)
    df.drop(['Date', 'Start datetime'], axis=1, inplace=True)

    retry_columns = sorted([col for col in df.columns if re.match(r'Backup \(Retry \d+\)', col)], key=lambda x: int(re.search(r'\d+', x).group()))
    sorted_columns = ['Week number', 'Day of week', 'Backup job', 'Backup'] + retry_columns + ['Status']

    df = df[sorted_columns]

    return merge_retry_rows(df)