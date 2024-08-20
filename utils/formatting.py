from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.formatting.rule import FormulaRule
import locale

def merge_cells(ws):
    cell_value = ws['B1']
    vm_list = [ws['C1'].value]
    merge_start = 2
    for row in range(2, ws.max_row + 1):
        current_cell = ws[f'B{row}']
        current_vm = ws[f'C{row}'].value
        if current_cell.value == cell_value:
            if current_vm not in vm_list:
                current_cell.value = None
                vm_list.append(current_vm)
            else:
                vm_list = [current_vm]
                if row > merge_start:
                    ws.merge_cells(start_row=merge_start, start_column=1, end_row=row - 1, end_column=1)
                    ws.merge_cells(start_row=merge_start, start_column=current_cell.column, end_row=row - 1, end_column=current_cell.column)
                cell_value = current_cell.value
                merge_start = row
        else:
            vm_list = [current_vm]
            if row > merge_start:
                ws.merge_cells(start_row=merge_start, start_column=1, end_row=row - 1, end_column=1)
                ws.merge_cells(start_row=merge_start, start_column=current_cell.column, end_row=row - 1, end_column=current_cell.column)
            cell_value = current_cell.value
            merge_start = row
    if merge_start < ws.max_row:
        ws.merge_cells(start_row=merge_start, start_column=1, end_row=ws.max_row, end_column=1)
        ws.merge_cells(start_row=merge_start, start_column=current_cell.column, end_row=ws.max_row, end_column=current_cell.column)


def format_header(ws):
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=ws.max_column):
        for cell in col:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment


def format_borders(ws):
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True)


def format_backup():
    locale.setlocale(locale.LC_TIME, 'pl_PL.UTF-8')

    workbook = load_workbook('workbooks/Backup data overview.xlsx')
    ws_backups = workbook['Backup']
    ws_details = workbook['Backup - objects']
    ws_last = workbook['Last backup']
    ws_last_obj = workbook['Last backup - objects']

    merge_cells(ws_details)
    merge_cells(ws_last_obj)

    format_header(ws_backups)
    format_header(ws_details)
    format_header(ws_last)
    format_header(ws_last_obj)

    format_borders(ws_backups)
    format_borders(ws_details)
    format_borders(ws_last)
    format_borders(ws_last_obj)


    al = Alignment(horizontal="left", vertical="top")
    for row in ws_details.iter_rows(min_col=1, max_col=2, min_row=1, max_row=ws_details.max_row):
        for cell in row:
            cell.alignment = al

    al = Alignment(horizontal="left", vertical="top")
    for row in ws_last_obj.iter_rows(min_col=1, max_col=2, min_row=1, max_row=ws_details.max_row):
        for cell in row:
            cell.alignment = al


    green_fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
    red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
    orange_fill = PatternFill(start_color='FFE5CC', end_color='FFE5CC', fill_type='solid')

    success_rule = FormulaRule(formula=['$C2="Success"'], fill=green_fill)
    error_rule = FormulaRule(formula=['$C2="Error"'], fill=red_fill)
    warning_rule = FormulaRule(formula=['$C2="Warning"'], fill=orange_fill)

    ws_last.conditional_formatting.add('A2:O' + str(ws_last.max_row), success_rule)
    ws_last.conditional_formatting.add('A2:O' + str(ws_last.max_row), error_rule)
    ws_last.conditional_formatting.add('A2:O' + str(ws_last.max_row), warning_rule)

    success_rule = FormulaRule(formula=['$D2="Success"'], fill=green_fill)
    error_rule = FormulaRule(formula=['$D2="Error"'], fill=red_fill)
    warning_rule = FormulaRule(formula=['$D2="Warning"'], fill=orange_fill)

    ws_last_obj.conditional_formatting.add('C2:J' + str(ws_last_obj.max_row), success_rule)
    ws_last_obj.conditional_formatting.add('C2:J' + str(ws_last_obj.max_row), error_rule)
    ws_last_obj.conditional_formatting.add('C2:J' + str(ws_last_obj.max_row), warning_rule)

    max_row = ws_last_obj.max_row
    for row in range(ws_last_obj.max_row, 0, -1):
        if any(cell.value is not None for cell in ws_last_obj[row]):
            max_row = row
            break

    job_colors = []
    for row in ws_last.iter_rows(min_row=2, max_row=max_row, values_only=True):
        job, status = row[1], row[2]
        if status == "Error":
            job_colors.append((job, red_fill))
        elif status == "Warning":
            job_colors.append((job, orange_fill))
        elif status == "Success":
            job_colors.append((job, green_fill))

    i = 2
    max_row = ws_last_obj.max_row

    for job, color in job_colors:
        while i < max_row - 1:
            row = ws_last_obj[i]
            date_cell = row[0]
            job_cell = row[1]
            if job_cell.value == job:
                date_cell.fill = color
                job_cell.fill = color
                job_colors = job_colors[1:]
                break
            else:
                i += 1
        i += 1

    workbook.save('workbooks/Backup data overview.xlsx')


def format_backup_execution():
    workbook = load_workbook('workbooks/Backup data overview.xlsx')
    worksheet = workbook['Backup execution']

    green_fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
    red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
    orange_fill = PatternFill(start_color='FFE5CC', end_color='FFE5CC', fill_type='solid')

    success_rule = FormulaRule(formula=['$H2="Success"'], fill=green_fill)
    error_rule = FormulaRule(formula=['$H2="Error"'], fill=red_fill)
    warning_rule = FormulaRule(formula=['$H2="Warning"'], fill=orange_fill)

    worksheet.conditional_formatting.add('C2:H' + str(worksheet.max_row), success_rule)
    worksheet.conditional_formatting.add('C2:H' + str(worksheet.max_row), error_rule)
    worksheet.conditional_formatting.add('C2:H' + str(worksheet.max_row), warning_rule)

    cell_value = {'week': None, 'day': None}
    merge_start_day = 2
    
    for row in range(2, worksheet.max_row + 1):
        week_number_cell = worksheet[f'A{row}']
        day_of_week_cell = worksheet[f'B{row}']
        
        if week_number_cell.value == cell_value['week'] and day_of_week_cell.value == cell_value['day']:
            day_of_week_cell.value = None
        else:
            if row > merge_start_day:
                worksheet.merge_cells(start_row=merge_start_day, start_column=2, end_row=row - 1, end_column=2)
            
            cell_value = {
                'week': week_number_cell.value,
                'day': day_of_week_cell.value
            }
            merge_start_day = row

    if merge_start_day < worksheet.max_row:
        worksheet.merge_cells(start_row=merge_start_day, start_column=2, end_row=worksheet.max_row, end_column=2)

    cell_value = None
    merge_start = 2
    for row in range(2, worksheet.max_row + 1):
        current_cell = worksheet[f'A{row}']
        if current_cell.value == cell_value:
            current_cell.value = None
        else:
            if row > merge_start:
                worksheet.merge_cells(start_row=merge_start, start_column=current_cell.column, end_row=row - 1, end_column=current_cell.column)
            cell_value = current_cell.value
            merge_start = row
    if merge_start < worksheet.max_row:
        worksheet.merge_cells(start_row=merge_start, start_column=current_cell.column, end_row=worksheet.max_row, end_column=current_cell.column)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col in worksheet.iter_cols(min_row=1, max_row=1, min_col=1, max_col=worksheet.max_column):
        for cell in col:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True)

    al = Alignment(horizontal="left", vertical="top")
    for row in worksheet.iter_rows(min_col=1, max_col=1, min_row=1, max_row=worksheet.max_row):
        for cell in row:
            cell.alignment = al

    for row in worksheet.iter_rows(min_col=2, max_col=2, min_row=1, max_row=worksheet.max_row):
        for cell in row:
            cell.alignment = al

    workbook.save('workbooks/Backup data overview.xlsx')