import pandas as pd
from openpyxl import load_workbook, Workbook
from utils.formatting import format_backup, format_execution
from utils.stats import stats_excel
import shutil
import os


def adjust_column_widths(writer, dataframe, sheet_name):
    worksheet = writer.sheets[sheet_name]
    for idx, col in enumerate(dataframe.columns):
        max_len = max(dataframe[col].astype(str).map(len).max(), len(col)) + 2
        worksheet.set_column(idx, idx, max_len)


def create_excels(backup_df, obj_df, last_backup_df, last_obj_df, execution_df, summary_df, summary_recent_df, largest_backups_df, smallest_backups_df, details_df, merged_counts_df):
    output_folder = 'workbooks'
    output_path = os.path.join(output_folder, 'Backup data overview.xlsx')
    
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    with pd.ExcelWriter(output_path, engine="xlsxwriter", date_format="yyyy-mm-dd") as writer:
        backup_df.to_excel(writer, sheet_name='Backup', index=False)
        obj_df.to_excel(writer, sheet_name='Backup - objects', index=False)
        last_backup_df.to_excel(writer, sheet_name='Last backup', index=False)
        last_obj_df.to_excel(writer, sheet_name='Last backup - objects', index=False)
        execution_df.to_excel(writer, sheet_name='Backup execution', index=False)

        adjust_column_widths(writer, backup_df, 'Backup')
        adjust_column_widths(writer, obj_df, 'Backup - objects')
        adjust_column_widths(writer, last_backup_df, 'Last backup')
        adjust_column_widths(writer, last_obj_df, 'Last backup - objects')
        adjust_column_widths(writer, execution_df, 'Backup execution')

    stats_excel(summary_df, summary_recent_df, largest_backups_df, smallest_backups_df, details_df, merged_counts_df)
    format_backup()
    format_execution()

    original_file = "workbooks/Backup data overview.xlsx"
    workbook = load_workbook(original_file)

    for sheet_name in workbook.sheetnames:
        new_file_name = f"workbooks/{sheet_name}.xlsx"
        shutil.copyfile(original_file, new_file_name)
        
        new_workbook = load_workbook(new_file_name)
        
        for name in new_workbook.sheetnames:
            if name != sheet_name:
                std = new_workbook[name]
                new_workbook.remove(std)
        
        new_workbook.save(new_file_name)