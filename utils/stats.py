import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from utils.data_processing import process_data


def generate_summary(df):
    avg_duration = df['Duration'].mean()
    avg_duration_str = str(avg_duration.components.hours).zfill(2) + ":" + \
        str(avg_duration.components.minutes).zfill(2) + ":" + \
        str(avg_duration.components.seconds).zfill(2)

    return {
        'Total Backups': len(df),
        'Successful Backups': len(df[df['Status'] == 'Success']),
        'Backups with Warnings': len(df[df['Status'] == 'Warning']),
        'Failed Backups': len(df[df['Status'] == 'Error']),
        'Machines with Failed Backups': df['Error'].sum(),
        'Average Backup Size (GB)': df['Backup Size (GB)'].mean(),
        'Average Backup Duration': avg_duration_str,
        'Average Speed (GB/min)': ((df['Data Read (GB)'] + df['Transferred (GB)']) / (df['Duration'].dt.total_seconds() / 60)).mean(),
        'Average Compression Ratio': df['Compression'].mean(),
        'Average Dedupe Ratio': df['Dedupe'].mean()
    }


def add_summary_to_sheet(sheet, df, start_row, start_col):
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
        for c_idx, value in enumerate(row, start_col):
            sheet.cell(row=r_idx, column=c_idx, value=value)


def format_headers(sheet, loc):
    sheet.merge_cells(start_row=loc[0], start_column=loc[1], end_row=loc[0], end_column=loc[2])
    cell = sheet.cell(row=loc[0], column=loc[1])
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def format_col_titles(sheet, loc):
    for col in range(loc[1], loc[2] + 1):
        cell = sheet.cell(row=loc[0], column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="A9C3E8", end_color="A9C3E8", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")


def format_borders(sheet, start_row, end_row, start_col, end_col):
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.border = thin_border


def auto_adjust_column_widths(sheet):
    for column_cells in sheet.columns:
        max_length = 0
        column = None
        for cell in column_cells:
            if cell.value:
                try:
                    column = cell.column_letter
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                except AttributeError:
                    continue
        if column:
            adjusted_width = max_length + 2
            sheet.column_dimensions[column].width = adjusted_width


def stats(backup_df, obj_df, last_backup_df, last_obj_df):
    backup_df, obj_df, last_backup_df, last_obj_df = process_data(backup_df, obj_df, last_backup_df, last_obj_df)

    general_summary = generate_summary(backup_df)
    summary_df = pd.DataFrame(list(general_summary.items()), columns=['Metric', 'Value'])

    general_summary_for_recent_backups = generate_summary(last_backup_df)
    summary_recent_df = pd.DataFrame(list(general_summary_for_recent_backups.items()), columns=['Metric', 'Value'])

    df_grouped = last_obj_df.groupby('Object').last().reset_index()
    obj_count = obj_df.groupby('Object').size().reset_index(name='Count')

    df_sorted = df_grouped.sort_values(by='Object').reset_index(drop=True)
    obj_count_sorted = obj_count.sort_values(by='Object').reset_index(drop=True)

    df_sorted['Datetime'] = df_sorted.apply(lambda row: datetime.combine(pd.to_datetime(row['Date']), pd.to_datetime(row['Start Time'], format='%H:%M:%S').time()), axis=1)

    details = {
        'Machine': df_sorted['Object'],
        'Last Backup Status': df_sorted['Status'],
        'Total Backups': obj_count_sorted['Count'],
        'Last Backup Date and Time': df_sorted['Datetime']
    }

    df_details = pd.DataFrame(details)

    total_counts = obj_df.groupby('Object').size().reset_index(name='Backup Count')
    error_counts = obj_df[obj_df['Status'] == 'Error'].groupby('Object').size().reset_index(name='Error Count')
    merged_counts = pd.merge(total_counts, error_counts, on='Object', how='left')
    merged_counts['Error Count'] = merged_counts['Error Count'].fillna(0)
    merged_counts['Error Rate'] = merged_counts['Error Count'] / merged_counts['Backup Count']
    merged_counts.rename(columns={'Object': 'Machine'}, inplace=True)
    merged_counts = merged_counts[['Machine', 'Error Rate']].sort_values(by='Error Rate', ascending=False)

    largest_backups = backup_df.nlargest(3, 'Backup Size (GB)')[['Backup Job', 'Backup Size (GB)']]
    df_no_error = backup_df[backup_df['Status'] != 'Error']
    smallest_backups = df_no_error.nsmallest(3, 'Backup Size (GB)')[['Backup Job', 'Backup Size (GB)']]

    return summary_df, summary_recent_df, largest_backups, smallest_backups, df_details, merged_counts


def stats_excel(summary_df, summary_recent_df, largest_backups_df, smallest_backups_df, details_df, merged_counts_df):
    wb = load_workbook('workbooks/Backup data overview.xlsx')
    if 'Summary' in wb.sheetnames:
        std = wb['Summary']
        wb.remove(std)
    ws = wb.create_sheet(title='Summary')

    ws['A1'] = "General Summary of Backups"
    add_summary_to_sheet(ws, summary_df, 2, 1)
    ws['D1'] = "General Summary of Recent Backups"
    add_summary_to_sheet(ws, summary_recent_df, 2, 4)
    ws[f'A{len(summary_df) + 4}'] = "Largest Backups"
    add_summary_to_sheet(ws, largest_backups_df, len(summary_df) + 5, 1)
    ws[f'D{len(summary_df) + 4}'] = "Smallest Backups"
    add_summary_to_sheet(ws, smallest_backups_df, len(summary_df) + 5, 4)
    ws[f'A{len(summary_df) + len(largest_backups_df) + 7}'] = "Machine Backup Summary"
    add_summary_to_sheet(ws, details_df, len(summary_df) + len(largest_backups_df) + 8, 1)
    ws[f'A{len(summary_df) + len(largest_backups_df) + len(details_df) + 10}'] = "Machine Backup Error Rate"
    add_summary_to_sheet(ws, merged_counts_df, len(summary_df) + len(largest_backups_df) + len(details_df) + 11, 1)

    header_locs = [[1, 1, 2], [1, 4, 5], [14, 1, 2], [14, 4, 5], [20, 1, 4], [len(details_df) + 23, 1, 2]]
    col_locs = [[2, 1, 2], [2, 4, 5], [15, 1, 2], [15, 4, 5], [21, 1, 4], [len(details_df) + 24, 1, 2]]
    border_locs = [
        (1, 12, 1, 2), (1, 12, 4, 5),
        (14, 18, 1, 2), (14, 18, 4, 5),
        (20, len(details_df) + 21, 1, 4),
        (len(details_df) + 23, 2 * len(details_df) + 24, 1, 2)
    ]

    for loc in header_locs:
        format_headers(ws, loc)
    for loc in col_locs:
        format_col_titles(ws, loc)
    for loc in border_locs:
        format_borders(ws, *loc)

    auto_adjust_column_widths(ws)

    wb.save('workbooks/Backup data overview.xlsx')
