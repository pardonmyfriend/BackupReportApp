import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import locale
from utils.backup_loader import report_summary, get_last_backups, get_job_objects, combine
from utils.execution_loader import get_backup_execution, merge_retry_rows, combine_exec


def load_data(files):
    backup_list = []
    obj_list = []
    execution_list = []
    
    for file in files:
        workbook = load_workbook(file)
        sheet_names = workbook.sheetnames

        for sheet_name in sheet_names:
            sheet = workbook[sheet_name]
            max_row = sheet.max_row
            cell_value = sheet.cell(row=max_row, column=1).value
            if cell_value and isinstance(cell_value, str) and "Veeam Backup & Replication" in cell_value:
                backup, obj = report_summary(sheet)
                backup_list.append(backup)
                obj_list.append(obj)
                execution_list.append(get_backup_execution(sheet))

    backup_df = combine(backup_list)
    obj_df = combine(obj_list)
    execution_df = merge_retry_rows(combine_exec(execution_list))

    last_backup_df, last_obj_df = get_last_backups(backup_df, obj_df)
    job_obj = get_job_objects(backup_df, obj_df)

    st.session_state['uploaded_backup'] = backup_df
    st.session_state['uploaded_obj'] = obj_df
    st.session_state['uploaded_execution'] = execution_df

    st.session_state['backup'] = backup_df
    st.session_state['obj'] = obj_df
    st.session_state['execution'] = execution_df
    st.session_state['last_backup'] = last_backup_df
    st.session_state['last_obj'] = last_obj_df
    st.session_state['job_obj'] = job_obj

    st.session_state['min_date'] = backup_df.loc[0, 'Date']
    st.session_state['max_date'] = backup_df.iloc[-1]['Date']
    st.session_state['year'] = pd.to_datetime(backup_df.loc[0, 'Date']).year


locale.setlocale(locale.LC_TIME, 'en_US')

st.header("Upload file")

if 'file_reset' in st.session_state and st.session_state['file_reset']:
    st.session_state.clear()
    st.success("Data has been reset. You can now upload a new file.")

if 'uploaded_backup' in st.session_state and 'uploaded_obj' in st.session_state:
    if 'file_just_loaded' in st.session_state and st.session_state['file_just_loaded']:
        st.success(":material/task_alt: File successfully loaded! Review the data below.")

    if 'file_just_loaded' in st.session_state and not st.session_state['file_just_loaded']:
        st.success(":material/task_alt: You've already loaded your file! Review the data below.")

    st.session_state['file_just_loaded'] = False

    backup_df = st.session_state['uploaded_backup']
    obj_df = st.session_state['uploaded_obj']

    tab1, tab2 = st.tabs(['Backup data', 'Detailed backup data by object'])
    tab1.dataframe(backup_df)
    tab2.dataframe(obj_df)

    if st.button('Reset data', use_container_width=True):
        st.session_state['file_reset'] = True
        st.switch_page("pages/upload_file.py")

    if st.button(f'Next step: Adjust parameters :material/arrow_forward_ios:', use_container_width=True, type='primary'):
        st.switch_page("pages/params.py")
else:
    files = st.file_uploader("Choose Excel file with Veeam report", type=["xlsx"], accept_multiple_files=True)

    if files:
        with st.spinner("Uploading..."):
            load_data(files)
        st.session_state['file_just_loaded'] = True
        st.switch_page("pages/upload_file.py")
